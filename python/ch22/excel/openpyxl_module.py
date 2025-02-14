#엑셀 파일 생성
from openpyxl import Workbook
wb=Workbook() #새로운 워크북 생성
sheet=wb.active #활성 시트 선택
sheet['A1']="Hello"
sheet["B1"]="World"

file=r"C:\rokey\ch22\excel\new_example.xlsx"
wb.save(file) #저장하기

#엑셀 파일 열기
from openpyxl import load_workbook
file=r"C:\rokey\ch22\excel\new_example.xlsx"
wb=load_workbook(file)
sheet=wb.active

#print(sheet["A1"])
#셀을 갖고와 출력 : <Cell 'Sheet'.A1>
print(sheet["A1"].value) #특정 셀 값 읽기
#출력 : Hello
print(sheet["B1"].value)

#여러 셀의 데이터 읽기
for row in sheet.iter_rows(min_row=1, max_col=3, max_row=5):
    #행 데이터가 들어옴 
    for cell in row:
        print(cell.value,end="\t")
    print()

#값 저장 및 수정
from datetime import datetime
sheet["A1"]=datetime(2025, 2, 10) #날짜 형태로 저장
sheet["B1"]=100
sheet["C1"]="=B1*100" #수식은 앞에 = 붙여야 함
file=r"C:\rokey\ch22\excel\new_example.xlsx"
wb.save(file)